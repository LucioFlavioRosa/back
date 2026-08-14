"""Acrescenta as 8 colunas `*_com_cts` à aba `subbacia-operacional`.

    universo_ligacoes_com_cts               ligacoes_atuais_com_cts
    universo_economias_com_cts              economias_atuais_com_cts
    universo_ligacoes_residencial_com_cts   ligacoes_atuais_residencial_com_cts
    universo_economias_residencial_com_cts  economias_atuais_residencial_com_cts

O QUE ELAS SIGNIFICAM

As colunas de ligação e economia da sub-bacia são o que pertence EXCLUSIVAMENTE a ela.
A CTS cobre uma área que se SOBREPÕE a essa, e a sobreposição é contada uma vez só, na
entidade que a atende em cada cenário:

    usar_cts=true    a CTS atende a sobreposição — ela está nos números da CTS
    usar_cts=false   a sub-bacia atende — e o total dela é `*_com_cts`

Antes disso o motor SOMAVA as duas linhas quando a CTS era desligada, e a ligação da
área sobreposta, que está nas duas, era contada duas vezes.

REGRA DE PREENCHIMENTO — e leia a ressalva, ela é o ponto

    com CTS pareada  ->  exclusiva + a quantidade da CTS
    sem CTS pareada  ->  a própria quantidade (sem coletor não há sobreposição)

O primeiro caso reproduz EXATAMENTE o que o motor já somava. Ou seja: rodar este
script **não muda resultado nenhum** hoje. Ele existe para mover a conta de dentro do
motor para dentro do dado, de modo que o dia em que o Databricks trouxer o valor
apurado — que vai ser MENOR onde houver sobreposição real — a mudança seja só de dado,
não de código.

Enquanto isso, `sobreposicao_origem` diz `derivado_soma` (o valor ainda embute a dupla
contagem) ou `sem_cts` (não havia o que sobrepor, e o valor é exato). Quem comparar uma
rodada sem CTS antes e depois da carga apurada precisa saber qual dos dois leu.

ORDEM: rode DEPOIS de `preencher_recorte_residencial.py` — as quatro colunas
residenciais são insumo das quatro `residencial_com_cts`.

Escrevo com openpyxl e não com `pandas.to_excel` porque este último reescreve o
arquivo inteiro e leva junto a formatação e as outras abas.

    python dev/preencher_sobreposicao_cts.py
    python dev/preencher_sobreposicao_cts.py --conferir
"""

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

PACOTE = Path(
    r"C:\Users\LúcioFláviodosSantos\OneDrive - Peers Consulting\Área de Trabalho"
    r"\aegea\Otimizador_CAPEX_v62_pacote_rev11\Otimizador_CAPEX_v62_pacote"
)
ARQ = PACOTE / "banco_dados_regional_v29_completo.xlsx"
ABA_SUB, ABA_CTS, ABA_PAR = "subbacia-operacional", "cts-operacional", "subbacia-cts"
ORIGEM = "sobreposicao_origem"

#: (coluna da sub-bacia, coluna equivalente na CTS, coluna consolidada a gravar)
PARES = [
    ("universo_ligacoes", "universo_ligacoes", "universo_ligacoes_com_cts"),
    ("ligacoes_atuais", "ligacoes_atuais", "ligacoes_atuais_com_cts"),
    ("universo_economias", "universo_economias", "universo_economias_com_cts"),
    ("economias_atuais", "economias_atuais", "economias_atuais_com_cts"),
    ("universo_ligacoes_residencial", "universo_ligacoes_residencial",
     "universo_ligacoes_residencial_com_cts"),
    ("ligacoes_atuais_residencial", "ligacoes_atuais_residencial",
     "ligacoes_atuais_residencial_com_cts"),
    ("universo_economias_residencial", "universo_economias_residencial",
     "universo_economias_residencial_com_cts"),
    ("economias_atuais_residencial", "economias_atuais_residencial",
     "economias_atuais_residencial_com_cts"),
]


def num(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def cabecalho(ws) -> dict[str, int]:
    return {str(c.value).strip(): i for i, c in enumerate(ws[1], start=1) if c.value}


def garantir(ws, cab: dict[str, int], nome: str) -> int:
    if nome in cab:
        return cab[nome]
    i = ws.max_column + 1
    ws.cell(1, i).value = nome
    cab[nome] = i
    return i


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--arquivo", default=str(ARQ))
    ap.add_argument("--conferir", action="store_true", help="relata sem gravar")
    a = ap.parse_args()

    arq = Path(a.arquivo)
    if not arq.exists():
        sys.exit(f"planilha nao encontrada: {arq}")

    wb = openpyxl.load_workbook(arq)
    for aba in (ABA_SUB, ABA_CTS, ABA_PAR):
        if aba not in wb.sheetnames:
            sys.exit(f"aba ausente: {aba}")

    # pareamento sub-bacia -> CTS
    wp = wb[ABA_PAR]
    cp = cabecalho(wp)
    ksub = cp.get("sub_bacia") or cp.get("subbacia")
    kcts = cp.get("cts") or cp.get("cts_id")
    par = {}
    for r in range(2, wp.max_row + 1):
        s, c = wp.cell(r, ksub).value, wp.cell(r, kcts).value
        if s and c:
            par[str(s).strip()] = str(c).strip()

    # quantidades da CTS, por id
    wk = wb[ABA_CTS]
    ck = cabecalho(wk)
    kid = ck.get("cts") or ck.get("cts_id")
    dados_cts = {}
    for r in range(2, wk.max_row + 1):
        cid = wk.cell(r, kid).value
        if not cid:
            continue
        dados_cts[str(cid).strip()] = {
            col_cts: num(wk.cell(r, ck[col_cts]).value) if col_cts in ck else None
            for _, col_cts, _ in PARES
        }

    ws = wb[ABA_SUB]
    cs = cabecalho(ws)
    ksb = cs.get("sub_bacia") or cs.get("subbacia")
    if not a.conferir:
        for _, _, cons in PARES:
            garantir(ws, cs, cons)
        garantir(ws, cs, ORIGEM)

    conta = {"derivado_soma": 0, "sem_cts": 0}
    for r in range(2, ws.max_row + 1):
        sb = ws.cell(r, ksb).value
        if not sb:
            continue
        cts = par.get(str(sb).strip())
        k = dados_cts.get(cts) if cts else None
        marca = "derivado_soma" if k else "sem_cts"
        conta[marca] += 1
        if a.conferir:
            continue
        for col_sub, col_cts, cons in PARES:
            base = num(ws.cell(r, cs[col_sub]).value) if col_sub in cs else None
            if base is None:
                continue
            extra = (k or {}).get(col_cts) or 0.0
            ws.cell(r, cs[cons]).value = int(round(base + extra))
        ws.cell(r, cs[ORIGEM]).value = marca

    print(
        f"  {ABA_SUB:<24} derivado_soma={conta['derivado_soma']:>5}  "
        f"sem_cts={conta['sem_cts']:>5}"
    )
    if a.conferir:
        print("(--conferir: nada gravado)")
        return

    bkp = arq.with_name(
        f"{arq.stem}.antes-da-sobreposicao-cts-{datetime.now():%Y%m%d-%H%M%S}{arq.suffix}"
    )
    shutil.copy2(arq, bkp)
    print(f"backup: {bkp.name}")
    wb.save(arq)
    print(f"gravado: {arq.name}")


if __name__ == "__main__":
    main()
