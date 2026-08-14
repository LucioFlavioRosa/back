"""Acrescenta o recorte RESIDENCIAL às abas `subbacia-operacional` e `cts-operacional`.

    universo_ligacoes_residencial    ligacoes_atuais_residencial
    universo_economias_residencial   economias_atuais_residencial

As quatro colunas existem no DDL, no banco, na API e na tela desde a mudança que
tirou o recorte industrial de cena — só não existiam na PLANILHA, que é a fonte de
`input.*`. Enquanto isso durar, quem recarregar o banco a partir dela repõe as
colunas vazias, e a rodada `COBERTURA_SO_RESIDENCIAL=True` volta a cair no total
(com alerta, mas cai).

REGRA DE PREENCHIMENTO

    sem indústria  -> residencial = total. A sub-bacia sem nenhum sinal industrial
                      é 100% residencial, e isso é afirmação, não chute.
    com indústria  -> residencial = total − parcela industrial, coluna a coluna
                      para ligações; para ECONOMIAS, pela proporção das ligações
                      (a planilha nunca teve economias industriais).

SOBRE ESSES VALORES — leia antes de confiar neles

Eles NÃO são medição. São derivação a partir das colunas industriais que a planilha
ainda carrega, e valem até o Databricks passar a exportar as quatro colunas
apuradas. É a mesma natureza do que `preencher_vazao_industrial.py` fazia — e a
mesma cautela: a coluna de controle `residencial_origem` fica ao lado dizendo
`derivado` ou `sem_industria`, para ninguém descobrir isso tarde demais.

Diferença importante para aquele script: aqui não há sorteio. A derivação é
aritmética sobre dado real, então rodar de novo dá o mesmo resultado sem precisar de
semente — e o dia em que o dado apurado chegar, ele simplesmente substitui.

AS COLUNAS INDUSTRIAIS CONTINUAM NA PLANILHA, e é de propósito. Elas não vão mais
para o banco (o DDL as removeu), mas são a ÚNICA evidência de onde os valores
derivados saíram. Apagá-las agora tornaria a derivação inauditável. Elas saem quando
a origem apurada chegar — junto com este script.

Escrevo com openpyxl e não com `pandas.to_excel` porque este último reescreve o
arquivo inteiro e leva junto a formatação e as outras abas.

    python dev/preencher_recorte_residencial.py                 # a planilha do pacote
    python dev/preencher_recorte_residencial.py --arquivo X.xlsx
    python dev/preencher_recorte_residencial.py --conferir      # só relata, não grava
"""

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

PACOTE = Path(
    r"C:\Users\LúcioFláviodosSantos\OneDrive - Peers Consulting\Área de Trabalho"
    r"\aegea\Otimizador_CAPEX_v62_pacote_rev11\Otimizador_CAPEX_v62_pacote"
)
ARQ = PACOTE / "banco_dados_regional_v29_completo.xlsx"
ABAS = ("subbacia-operacional", "cts-operacional")
ORIGEM = "residencial_origem"

#: total -> residencial -> parcela industrial que sai do total.
#: Economias não têm coluna industrial na planilha: derivam pela proporção das
#: ligações, que é a mesma regra que a migração 02 do banco usou. Manter as duas
#: iguais é o que faz o banco recarregado bater com o banco migrado.
LIGACOES = [
    ("universo_ligacoes", "universo_ligacoes_residencial", "universo_ligacoes_industrial"),
    ("ligacoes_atuais", "ligacoes_atuais_residencial", "ligacoes_atuais_industrial"),
]
ECONOMIAS = [
    ("universo_economias", "universo_economias_residencial", "universo_ligacoes"),
    ("economias_atuais", "economias_atuais_residencial", "ligacoes_atuais"),
]

#: Qualquer sinal destes significa "há indústria nesta linha".
SINAIS = [
    "universo_ligacoes_industrial",
    "ligacoes_atuais_industrial",
    "receita_faturada_industrial",
    "receita_arrecadada_industrial",
]


def num(v):
    """Número da célula, ou None. Célula vazia não é zero — zero é uma afirmação."""
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


def coluna(ws, nome: str) -> int | None:
    for i, c in enumerate(ws[1], start=1):
        if str(c.value).strip() == nome:
            return i
    return None


def garantir(ws, nome: str) -> int:
    """Índice da coluna, criando-a no fim do cabeçalho se ainda não existir."""
    i = coluna(ws, nome)
    if i:
        return i
    i = ws.max_column + 1
    ws.cell(1, i).value = nome
    return i


def processar(ws, conferir: bool) -> dict[str, int]:
    idx = {n: coluna(ws, n) for grupo in (LIGACOES, ECONOMIAS) for n, _, _ in grupo}
    for n in SINAIS:
        idx[n] = coluna(ws, n)
    if not conferir:
        for grupo in (LIGACOES, ECONOMIAS):
            for _, res, _ in grupo:
                idx[res] = garantir(ws, res)
        idx[ORIGEM] = garantir(ws, ORIGEM)

    conta = {"derivado": 0, "sem_industria": 0, "sem_total": 0}
    for r in range(2, ws.max_row + 1):
        tem_industria = any(
            idx.get(s) and (num(ws.cell(r, idx[s]).value) or 0) > 0 for s in SINAIS
        )
        marca = "derivado" if tem_industria else "sem_industria"

        for tot, res, ind in LIGACOES:
            vt = num(ws.cell(r, idx[tot]).value) if idx.get(tot) else None
            if vt is None:
                conta["sem_total"] += 1
                continue
            vi = num(ws.cell(r, idx[ind]).value) if idx.get(ind) else None
            if not conferir:
                ws.cell(r, idx[res]).value = int(round(max(0.0, vt - (vi or 0.0))))

        # Economias acompanham a proporção das ligações da própria linha.
        for tot, res, base_lig in ECONOMIAS:
            ve = num(ws.cell(r, idx[tot]).value) if idx.get(tot) else None
            vl = num(ws.cell(r, idx[base_lig]).value) if idx.get(base_lig) else None
            if ve is None or not vl:
                continue
            lig_res_col = next(rr for tt, rr, _ in LIGACOES if tt == base_lig)
            vlr = (
                num(ws.cell(r, idx[lig_res_col]).value)
                if (not conferir and idx.get(lig_res_col))
                else None
            )
            if vlr is None:
                continue
            if not conferir:
                ws.cell(r, idx[res]).value = int(round(ve * vlr / vl))

        if not conferir:
            ws.cell(r, idx[ORIGEM]).value = marca
        conta[marca] += 1
    return conta


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--arquivo", default=str(ARQ))
    ap.add_argument("--conferir", action="store_true", help="relata sem gravar")
    a = ap.parse_args()

    arq = Path(a.arquivo)
    if not arq.exists():
        sys.exit(f"planilha nao encontrada: {arq}")

    if not a.conferir:
        # Backup ao lado, com carimbo — o mesmo hábito do script da vazão. Uma
        # planilha de 4.850 linhas reescrita sem volta é um dia de trabalho perdido.
        bkp = arq.with_name(
            f"{arq.stem}.antes-do-recorte-residencial-"
            f"{datetime.now():%Y%m%d-%H%M%S}{arq.suffix}"
        )
        shutil.copy2(arq, bkp)
        print(f"backup: {bkp.name}")

    wb = openpyxl.load_workbook(arq)
    for aba in ABAS:
        if aba not in wb.sheetnames:
            print(f"  {aba:<24} (ausente)")
            continue
        c = processar(wb[aba], a.conferir)
        print(
            f"  {aba:<24} derivado={c['derivado']:>5}  "
            f"sem_industria={c['sem_industria']:>5}  sem_total={c['sem_total']:>3}"
        )
    if a.conferir:
        print("(--conferir: nada gravado)")
        return
    wb.save(arq)
    print(f"gravado: {arq.name}")


if __name__ == "__main__":
    main()
