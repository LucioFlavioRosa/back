"""Acrescenta `vazao_contribuicao_industrial` a aba `subbacia-operacional`.

A coluna existe no DDL, no banco, na API e na tela — so nao existia na planilha,
que e a fonte. Por isso ela chegava NULL nas 4.850 linhas e nao havia de onde
preencher. A aba `cts-operacional` ja tem essa coluna (49 preenchidas): a de
sub-bacia e que ficou para tras.

REGRA DE PREENCHIMENTO
  sem industria   -> 0. Sao as sub-bacias sem NENHUM sinal industrial (nem
                    universo de ligacoes, nem ligacoes atuais, nem receita).
                    Zero aqui e afirmacao, nao chute: se nao ha industria, a
                    parcela industrial da vazao e zero mesmo.
  com industria   -> fracao SORTEADA entre 5% e 12% da `vazao_contribuicao`.

SOBRE OS VALORES SORTEADOS — leia antes de confiar neles
  Eles NAO sao medicao. Sao um preenchimento plausivel para destravar a analise
  so-residencial (`INCLUIR_INDUSTRIAL=False`), que hoje nao roda por falta do
  campo. Quem for usar o resultado dessa analise precisa saber que a parcela
  industrial e sintetica. A coluna de controle `vazao_industrial_origem` fica ao
  lado dizendo `sorteado` ou `sem_industria`, exatamente para que ninguem
  descubra isso tarde demais.

  A semente e FIXA (20260807). Rodar de novo da os mesmos numeros: um dado
  sintetico que muda a cada execucao seria pior que nenhum, porque duas rodadas
  do mesmo cadastro divergiriam sem explicacao.

  Escrevo com openpyxl e nao com `pandas.to_excel` porque este ultimo reescreve o
  arquivo inteiro e leva junto formatacao e as outras 14 abas.
"""

import random
import shutil
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

PACOTE = Path(
    r"C:\Users\LúcioFláviodosSantos\OneDrive - Peers Consulting\Área de Trabalho"
    r"\aegea\Otimizador_CAPEX_v62_pacote_rev11\Otimizador_CAPEX_v62_pacote"
)
ARQ = PACOTE / "banco_dados_regional_v29_completo.xlsx"
ABA = "subbacia-operacional"
COL = "vazao_contribuicao_industrial"
ORIGEM = "vazao_industrial_origem"
SEMENTE = 20260807
PISO, TETO = 0.05, 0.12

#: Qualquer sinal destes significa "ha industria nesta sub-bacia".
SINAIS = [
    "universo_ligacoes_industrial",
    "ligacoes_atuais_industrial",
    "receita_faturada_industrial",
    "receita_arrecadada_industrial",
]


def main() -> None:
    if not ARQ.exists():
        sys.exit(f"nao achei {ARQ}")

    backup = ARQ.with_name(
        f"{ARQ.stem}.antes-de-{COL}-{datetime.now():%Y%m%d-%H%M%S}{ARQ.suffix}"
    )
    shutil.copy2(ARQ, backup)
    print(f"copia de seguranca: {backup.name}")

    wb = openpyxl.load_workbook(ARQ)
    ws = wb[ABA]
    cab = {c.value: c.column for c in ws[1] if c.value}

    if COL in cab:
        sys.exit(f"a coluna `{COL}` ja existe na aba {ABA} — nada a fazer")
    for s in SINAIS + ["vazao_contribuicao"]:
        if s not in cab:
            sys.exit(f"coluna esperada ausente: {s}")

    c_vaz = cab["vazao_contribuicao"]
    c_sinais = [cab[s] for s in SINAIS]
    c_novo = ws.max_column + 1
    c_origem = c_novo + 1
    ws.cell(row=1, column=c_novo, value=COL)
    ws.cell(row=1, column=c_origem, value=ORIGEM)

    rnd = random.Random(SEMENTE)
    com = sem = sem_vazao = 0

    for lin in range(2, ws.max_row + 1):
        tem_industria = any(
            (v := ws.cell(row=lin, column=c).value) is not None and v != 0 for c in c_sinais
        )
        if not tem_industria:
            ws.cell(row=lin, column=c_novo, value=0)
            ws.cell(row=lin, column=c_origem, value="sem_industria")
            sem += 1
            continue

        vaz = ws.cell(row=lin, column=c_vaz).value
        if vaz is None:
            # Nao inventa parcela de um total que nao existe: sem a vazao total,
            # 5-12% dela nao e uma conta que se possa fazer.
            ws.cell(row=lin, column=c_origem, value="sem_vazao_total")
            sem_vazao += 1
            continue

        ws.cell(row=lin, column=c_novo, value=round(float(vaz) * rnd.uniform(PISO, TETO), 2))
        ws.cell(row=lin, column=c_origem, value="sorteado")
        com += 1

    wb.save(ARQ)
    print(f"\naba {ABA}: {ws.max_row - 1} linhas")
    print(f"  sem industria (0)        {sem:>6}")
    print(f"  com industria (sorteado) {com:>6}   {PISO:.0%}-{TETO:.0%} da vazao, semente {SEMENTE}")
    if sem_vazao:
        print(f"  sem vazao total (vazio)  {sem_vazao:>6}   <- nao dava para calcular")
    print(f"\ncoluna `{ORIGEM}` ao lado registra a procedencia de cada valor.")


if __name__ == "__main__":
    main()
